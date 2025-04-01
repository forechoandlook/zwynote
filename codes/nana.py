

from collections import defaultdict
import pandas as pd
import sys
import json
import numpy as np
from tqdm import tqdm

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

df_path      = sys.argv[1]
json_path    = sys.argv[2]
output_path  = sys.argv[3]
name_info    = sys.argv[4]

df           = pd.read_csv(df_path)
profile_info = {}
index_keys   = ['Engine Id', "Core Id", "Cmd Id"]
value_keys   = ["Function Type", "Function Name", "Asic Cycle", "Data Type", "Start Cycle", "End Cycle"]
dtype_len    = {"f16":2, "f32":4, "i32":4, "si32":4, "bf16":2}
location     = json.load(open(json_path,'r'))

wb = Workbook()
ws = wb.active
ws.title = f"PCIE {name_info} Summary"

dma_keys     = ["DMA data size(B)", "Stall Cycle", "DDR Bandwidth(GB/s)", "L2M Bandwidth(GB/s)", "Direction", "src_nsize", "src_csize", "src_hsize", "src_wsize", "src_nstride", "src_cstride", "src_hstride", "src_wstride", "dst_nstride", "dst_cstride", "dst_hstride", "dst_wstride"]

#["DMA data size(B)", "Stall Cycle", "DDR Bandwidth(GB/s)", "L2M Bandwidth(GB/s)", "Direction"]

#["src_nsize", "src_csize", "src_hsize", "src_wsize", "src_nstride", "src_cstride", "src_hstride", "src_wstride"]
#["dst_nstride", "dst_cstride", "dst_hstride", "dst_wstride"]


GLOBAL_MEM_START = 2e12
L2_MEM_START     = 3.3e12
LOCAL_MEM_START  = 3.4e12


def write_to_sheet(ws, data, start_row, start_col, color_cells=None, color="FFFF00", default_width=15):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    if isinstance(data, pd.DataFrame):
        flag=False
        for r_idx, row in enumerate(dataframe_to_rows(data, index=True, header=True), start=start_row):
            if row == [None]:
                flag = True
                continue
            for c_idx, value in enumerate(row, start=start_col):
                cell = ws.cell(row=r_idx -1  if flag else r_idx, column=c_idx, value=value)
                if color_cells is not None and (r_idx-1-start_row, c_idx-start_col) in color_cells:
                    cell.fill = fill
                if default_width:
                    col_letter = ws.cell(row=start_row, column=c_idx).column_letter
                    ws.column_dimensions[col_letter].width = default_width
        return r_idx
    elif isinstance(data, list):
        for r_idx, line in enumerate(data, start=start_row):
            cell = ws.cell(row=r_idx, column=start_col, value=line)
            if color_cells is not None and (r_idx, start_col) in color_cells:
                cell.fill = fill
        return r_idx
    else:
        raise ValueError("Unsupported data type")

L2_SRAM_SIZE = 0x8000000

index_indices = [df.columns.get_loc(col) for col in index_keys]
value_indices = [df.columns.get_loc(col) for col in value_keys]

profile_info = {
    tuple(row[idx] for idx in index_indices): {value_keys[i]: row[val] for i, val in enumerate(value_indices)}
    for row in tqdm(df.itertuples(index=False, name=None))
}

# dma_profile_info = {
#     tuple(row[idx] for idx in index_indices): {dma_keys[i]: row[val] for i, val in enumerate(dma_keys)}
#     for row in tqdm(df.itertuples(index=False, name=None))
# }

def calc_single_op_time(op):
    tiu_b, dma_b  = op['tiu_dma_id(before)']
    tiu_a, dma_a  = op['tiu_dma_id(after)']
    if tiu_a == tiu_b and dma_a == dma_b:
        return 0, 0, 0, 0, 0

    total_start_cycle = 100000000000000
    total_end_cycle   = 0
    core_id           = op['core_id']
    dma_time          = 0
    tiu_time          = 0
    dma_sys_time      = 0
    tiu_sys_time      = 0
    for dma_id in range(dma_b+1, dma_a+1):
        key = (1, core_id, dma_id)
        if key not in profile_info:
            continue
        res = profile_info[key]
        total_start_cycle = min(total_start_cycle, res["Start Cycle"])
        total_end_cycle   = max(total_end_cycle, res['End Cycle'])
        if "sys" in res['Function Name'].lower():
            dma_sys_time += res['Asic Cycle']
            continue
        dma_time += res['Asic Cycle']

    for tiu_id in range(tiu_b+1, tiu_a + 1):
        key = (0, core_id, tiu_id)
        if key not in profile_info:
            continue
        res = profile_info[key]
        total_start_cycle = min(total_start_cycle, res["Start Cycle"])
        total_end_cycle   = max(total_end_cycle, res['End Cycle'])
        if "sys" in res['Function Name'].lower():
            tiu_sys_time += res['Asic Cycle']
            continue
        tiu_time += res['Asic Cycle']

    if total_end_cycle > 10000000000:
        breakpoint()

    total_time = total_end_cycle - total_start_cycle
    return tiu_time, tiu_sys_time, dma_time, dma_sys_time, total_time


op_set = defaultdict(lambda : defaultdict(list))
layer_set = set()
layer_id = 0

group_info = defaultdict(list)
group_idx = 0
group_start = 0
for each_op in location:
    layer_name ="_".join([each_op['opcode'], *[ i['name'] if "name" in i else "" for i in each_op['operands']] ])
    if layer_name not in layer_set:
        layer_id += 1
        layer_set.add(layer_name)
    each_op["layer_id"] = layer_id
    if each_op["core_id"] == 0:
        if each_op["is_local"]:
            if group_start:
                group_info[group_idx].append(each_op["opcode"])
            else:
                group_start = 1
                group_idx += 1
                group_info[group_idx].append(each_op["opcode"])
        else:
            group_start = 0

    tiu_time, tiu_sys_time, dma_time, dma_sys_time, total_time = calc_single_op_time(each_op)
    each_op["dma_time"]     = dma_time
    each_op["tiu_time"]     = tiu_time
    each_op["dma_sys_time"] = dma_sys_time
    each_op["tiu_sys_time"] = tiu_sys_time
    each_op["total_time"]   = total_time
    op_set[each_op["opcode"]][each_op['core_id']].append(each_op)

def get_op_shape_info(op) -> str:
    opinfos = ""
    if "Reshape" in op['opcode']:
        return opinfos
    for each in op["operands"]:
        if each == {}:
            opinfos += "None "
            continue
        opinfos += f"{each['memory_type']}"
    opinfos += " ===>>>> "
    for each in op["results"]:
        if each == {}:
            opinfos += "None "
            continue
        opinfos += f"{each['memory_type']}"
    return opinfos

def default_op_summary():
    default_dict = {}
    default_dict[f"local_count"]        = 0
    default_dict[f"total_count"]        = 0
    for i in range(8):
        default_dict[f"{i}_dma_time"]     = 0
        default_dict[f"{i}_tiu_time"]     = 0
        default_dict[f"{i}_dma_sys_time"] = 0
        default_dict[f"{i}_tiu_sys_time"] = 0
        default_dict[f"{i}_total_time"]   = 0
    return default_dict

op_summary = defaultdict(lambda: default_op_summary())

bmodel_annotation     = ["The following is bmodel pcie profile"]
bmodel_annotation.append("Column Description: coreidx_engine_time, such as 0_tiu_time")
bmodel_annotation.append("Since core_0 will take the longest time, we summarize the time of core_0 and callucate the time precent of each operation")
bmodel_annotation.append("    Time(cycle ns)      ")
bmodel_annotation.append(" local count 指的是 是local的个数  ")
bmodel_annotation.append(" total count 指的是 是总的算子个数  ")
# bmodel_annotation.append(" 注意 tile 可以消除  ")
# bmodel_annotation.append(" 目前根据final mlir所有的切分最多只有8份，可以平均到8个核，从而一个group只会运行一次！所以里面的load，store并没有很好和计算并行起来。  ")


bmodel_annotation_after = ["Attention:"]
bmodel_annotation_after.append("tiu_time: calculation time of tiu without SYS ( system.send ) time. ")
bmodel_annotation_after.append("dma_time: calculation time of tiu without DMA_sys ( sys_wait ) time. ")
bmodel_annotation_after.append("tiu_sys_time: only calculation time of tiu with SYS ( system.send ) time. ")
bmodel_annotation_after.append("dma_sys_time: only calculation time of tiu with DMA_sys ( sys_wait ) time. ")
bmodel_annotation_after.append("Extract the time from PerfAI_output.csv file generated by mlir.")
bmodel_annotation_after.append("Since some commands may be not recorded in PMU. The data may not be very accurate.")

core_0_total_time = 0
pcie_op_info = []
for op_name in op_set.keys():
    for core_id in op_set[op_name].keys():
        for res in op_set[op_name][core_id]:
            op_summary[op_name][f"{core_id}_dma_time"]       += res['dma_time']
            op_summary[op_name][f"{core_id}_tiu_time"]       += res['tiu_time']
            op_summary[op_name][f"{core_id}_total_time"]     += res['total_time']
            if core_id == 0:
                op_summary[op_name]['local_count'] += res["is_local"]
                op_summary[op_name]['total_count']       += 1
                core_0_total_time += res['total_time']
                pcie_op_info.append({"name": op_name, "is_local": bool(res["is_local"]), "dma_time": res['dma_time'], "tiu_time": res['tiu_time'],
                                    "dma_sys_time": res['dma_sys_time'], "tiu_sys_time": res['tiu_sys_time'], "info": get_op_shape_info(res),
                                    "total_time": res['total_time']})
            op_summary[op_name][f"{core_id}_dma_sys_time"]   += res['dma_sys_time']
            op_summary[op_name][f"{core_id}_tiu_sys_time"]   += res['tiu_sys_time']

new_pd = pd.DataFrame(op_summary).T

# 保留两位小数
new_pd.insert((new_pd.columns.get_loc("0_total_time")+1), "precent", round(new_pd["0_total_time"]/core_0_total_time, 4) * 100)
new_pd = new_pd.sort_values(by="precent", ascending=False)
# add summary
new_pd.loc["Summary"] = new_pd.select_dtypes(include=[float, int]).sum()
# color precent > 8
new_pd = new_pd.reset_index(drop=False)
color_cells = set((1, new_pd.columns.get_loc("precent") + 1))
for idx, row in new_pd.iterrows():
    if row['precent'] > 8:
        color_cells.add((idx + 1, new_pd.columns.get_loc("precent") + 1))
    if row['local_count'] > 0:
        color_cells.add((idx + 1, new_pd.columns.get_loc("local_count") + 1))

last_row = write_to_sheet(ws, bmodel_annotation, 3, 3)
last_row = write_to_sheet(ws, new_pd, last_row + 3, 3, color_cells)
last_row = write_to_sheet(ws, bmodel_annotation_after, last_row + 2, 3)

core_0_df = pd.DataFrame(pcie_op_info)
core_0_df["Precent"] = round(core_0_df["total_time"] / core_0_total_time * 100, 2)
core_0_df = core_0_df.sort_values(by="Precent", ascending=False)

# # summary
core_0_df.loc["Summary"] = core_0_df.select_dtypes(include=[float, int]).sum()
core_0_annotation = []
core_0_annotation.append("下面是core0上的每一个op的时间。出于core0肯定是每一个op都有，且是多core最长的时间。")
core_0_annotation.append("info 是op的输入和输出")
last_row = write_to_sheet(ws, core_0_annotation, last_row + 3, 3)
last_row = write_to_sheet(ws, core_0_df, last_row + 3, 3)


# group info
filter_ops = ["tpu.Load","tpu.Store"]
filter_other = lambda x: [i for i in x if i not in filter_ops]

ws = wb.create_sheet(title=f"Group INFO {name_info}")

annotation = [ "从location json里面提取的group信息", "", "" ]
annotation += [ " ".join( filter_other(group_info[k]) ) for k in group_info.keys()]
last_row = write_to_sheet(ws, annotation, 3, 3)

def parse_shape_dtype(memory_type):
    # <1xf32>
    mem_str = memory_type[1:-1]
    mem_split = mem_str.split("x")
    dtype = mem_split[-1]
    shape = mem_split[:-1]
    return [int(i) for i in shape], dtype

def shape_dtype_size(memory_type):
    shape, dtype = parse_shape_dtype(memory_type)
    return np.prod(shape) * dtype_len[dtype]

def dma_each_op(op):
    inputs = op['operands']
    outputs= op['results']
    dms = 0
    for each in inputs:
        if each == {}:
            continue
        if "buffer" in each['name']:
            continue
        dms += shape_dtype_size(each['memory_type'])

    for each in outputs:
        if each == {}:
            continue
        if "buffer" in each['name']:
            continue
        dms += shape_dtype_size(each['memory_type'])
    return dms


def get_idx_input(op, idx):
    return op['operands'][idx]

def get_idx_input_shape(op, idx):
    input_ele = op['operands'][idx]
    mem_type = input_ele['memory_type']
    shape, dtype = parse_shape_dtype(mem_type)
    return shape

def get_idx_output_shape(op, idx):
    input_ele = op['results'][idx]
    mem_type = input_ele['memory_type']
    shape, dtype = parse_shape_dtype(mem_type)
    return shape

op_factory = {}
def warp_calc(fn):
    op_factory[fn.__name__] = fn
    def _warp(*args, **kwargs):
        return fn(*args, **kwargs)
    return _warp

@warp_calc
def Reshape_op_dma_tiu(op):
    return 0, 0, 0

@warp_calc
def Cast_op_dma_tiu(op):
    input1_shape = get_idx_input_shape(op, 0)
    return dma_each_op(op), 0, np.prod(input1_shape) * 1

cell = lambda x, b: (x+b-1)//b

@warp_calc
def Conv2D_op_dma_tiu(op):
    input1_shape = get_idx_input_shape(op, 0)
    input2_shape = get_idx_input_shape(op, 1)
    output1_shape = get_idx_output_shape(op,0)
    _,cout,hout,wout = output1_shape
    n,c,h,w = input1_shape
    _,oc,inner,lane = input2_shape
    khkw = inner / cell(c, 32)
    cube = 2 * n * c * oc * hout * wout * khkw
    return dma_each_op(op), cube, 0

@warp_calc
def ConvBnReluBwdOp_op_dma_tiu(op):
    return 0, 0, 0


# MeanRstd
@warp_calc
def MeanRstd_op_dma_tiu(op):
    input1_shape = get_idx_input_shape(op, 0)
    return dma_each_op(op), 0, np.prod(input1_shape) * 10

@warp_calc
def AddConst_op_dma_tiu(op):
    input1_shape = get_idx_input_shape(op, 0)
    return dma_each_op(op), 0, np.prod(input1_shape) * 1

@warp_calc
def Scale_op_dma_tiu(op):
    input1_shape = get_idx_input_shape(op, 0)
    return dma_each_op(op), 0, np.prod(input1_shape) * 10

@warp_calc
def MaxPoolWithMask_op_dma_tiu(op):
    input1_shape = get_idx_input_shape(op, 0)
    output1_shape = get_idx_output_shape(op,0)
    n,c,h,w = input1_shape
    kh = kw = 3
    return dma_each_op(op), 0, n*c*h*w*kh*kw*2

@warp_calc
def Add_op_dma_tiu(op):
    output_shape = get_idx_output_shape(op,0)
    return dma_each_op(op), 0, np.prod(output_shape) * 1

@warp_calc
def Pool2D_op_dma_tiu(op):
    kh = kw = 7
    output_shape = get_idx_output_shape(op,0)
    n,c,h,w = output_shape
    return dma_each_op(op), 0, kw*kh*n*c*h*w*2

@warp_calc
def Permute_op_dma_tiu(op):
    return dma_each_op(op), 0, 0

@warp_calc
def CompareConst_op_dma_tiu(op):
    input_shape = get_idx_input_shape(op, 0)
    return dma_each_op(op), 0, np.prod(input_shape) * 2

@warp_calc
def Concat_op_dma_tiu(op):
    return dma_each_op(op), 0, 0

@warp_calc
def SwapDimInner_op_dma_tiu(op):
    return dma_each_op(op), 0, 0

@warp_calc
def WhereBnbwdOp_op_dma_tiu(op):
    output_shape = get_idx_output_shape(op, 0)
    return dma_each_op(op), 0, np.prod(output_shape) * 6

@warp_calc
def GroupNormTrain_op_dma_tiu(op):
    input1_shape = get_idx_input_shape(op, 0)
    return dma_each_op(op), 0, np.prod(input1_shape) * 10

@warp_calc
def LayerNormTrain_op_dma_tiu(op):
    input1_shape = get_idx_input_shape(op, 0)
    return dma_each_op(op), 0, np.prod(input1_shape) * 10

@warp_calc
def Slice_op_dma_tiu(op):
    return dma_each_op(op), 0, 0

@warp_calc
def Gather_op_dma_tiu(op):
    return dma_each_op(op), 0, 0

@warp_calc
def MatMul_op_dma_tiu(op):
    input1_shape = get_idx_input_shape(op, 0)
    input2_shape = get_idx_input_shape(op, 1)
    res_shape    = get_idx_output_shape(op, 0)
    has_bias = op['operands'][2] != {}
    if len(input1_shape) == 3:
        return dma_each_op(op), (24 if has_bias else 16) * np.prod(input1_shape) * np.prod(input2_shape) / input1_shape[-1] / input1_shape[0], 0
    # print(input1_shape, input2_shape, res_shape)
    return dma_each_op(op), (3 if has_bias else 2) * np.prod(input1_shape) * np.prod(input2_shape) / input1_shape[-1], 0

@warp_calc
def Softmax_op_dma_tiu(op):
    input1_shape = get_idx_input_shape(op, 0)
    return dma_each_op(op), 0, np.prod(input1_shape) * 6

@warp_calc
def BatchNormTrain_op_dma_tiu(op):
    input1_shape = get_idx_input_shape(op, 0)
    return dma_each_op(op), 0, np.prod(input1_shape) * 8

@warp_calc
def LayerNorm_op_dma_tiu(op):
    input1_shape = get_idx_input_shape(op, 0)
    return dma_each_op(op), 0, np.prod(input1_shape) * 8

@warp_calc
def Where_op_dma_tiu(op):
    input1_shape = get_idx_input_shape(op, 0)
    return dma_each_op(op), 0, np.prod(input1_shape)

@warp_calc
def Relu_op_dma_tiu(op):
    input1_shape = get_idx_input_shape(op, 0)
    return dma_each_op(op), 0, np.prod(input1_shape)

@warp_calc
def MaxPoolingIndicesBwd_op_dma_tiu(op):
    output_shape = get_idx_output_shape(op, 0)
    return dma_each_op(op), 0, np.prod(output_shape) * 9


@warp_calc
def GatherElements_op_dma_tiu(op):
    return dma_each_op(op),0,0

@warp_calc
def MulConst_op_dma_tiu(op):
    output_shape = get_idx_output_shape(op, 0)
    return dma_each_op(op), 0, np.prod(output_shape)

@warp_calc
def Div_op_dma_tiu(op):
    output_shape = get_idx_output_shape(op, 0)
    return dma_each_op(op),0, np.prod(output_shape)

# ScatterElements
@warp_calc
def ScatterElements_op_dma_tiu(op):
    return dma_each_op(op),0,0

@warp_calc
def Load_op_dma_tiu(op):
    return 0,0,0

@warp_calc
def Store_op_dma_tiu(op):
    return 0,0,0

# Mul
@warp_calc
def Mul_op_dma_tiu(op):
    output_shape = get_idx_output_shape(op, 0)
    return dma_each_op(op),0, np.prod(output_shape)

# Active
@warp_calc
def Active_op_dma_tiu(op):
    output_shape = get_idx_output_shape(op, 0)
    return dma_each_op(op),0, np.prod(output_shape) * 10

# Reduce
@warp_calc
def Reduce_op_dma_tiu(op):
    input1_shape = get_idx_input_shape(op, 0)
    return dma_each_op(op),0, np.prod(input1_shape) * 2

# Sub
@warp_calc
def Sub_op_dma_tiu(op):
    output_shape = get_idx_output_shape(op, 0)
    return dma_each_op(op),0, np.prod(output_shape)

# Tile
@warp_calc
def Tile_op_dma_tiu(op):
    return dma_each_op(op),0,0

# BatchNormBwd
@warp_calc
def BatchNormBwd_op_dma_tiu(op):
    input1_shape = get_idx_input_shape(op, 0)
    n,c,h,w = input1_shape
    return dma_each_op(op), 0, np.prod(input1_shape)*7 + 2*c

# Convbwd
@warp_calc
def Convbwd_op_dma_tiu(op):
    calc_grad_input  = op['results'][0] != {}
    calc_grad_weight = op['results'][1] != {}

    kernel_shape     = get_idx_input_shape(op, 2)
    oc,ic,kh,kw      = kernel_shape
    output_shape     = get_idx_input_shape(op, 0)
    input_shape      = get_idx_input_shape(op, 1)
    n,c,h,w          = input_shape
    _,_,hout,wout    = output_shape
    cube = 0
    vec  = 0
    if calc_grad_input:
        cube += n * oc * hout * wout * c * kh * kw * 2
    if calc_grad_weight:
        cube += n * oc * h * w * c * kh * kw * 2
    return dma_each_op(op), cube, vec

# Pad
@warp_calc
def Pad_op_dma_tiu(op):
    return dma_each_op(op), 0, 0

def calc_op_dma_ops(op):
    name = op['opcode']
    fn_name = name.replace("tpu.","") + "_op_dma_tiu"
    return op_factory[fn_name](op)

CONFIG = {
    "dma" : 548e3,
    "cube": 128e6,
    "vec" : 32e6
}

def calc_time(dma, cube, vec):
    dma_time   = dma/CONFIG['dma']
    cube_time  = cube/CONFIG['cube']
    vec_time   = vec/CONFIG['vec']
    total_time = max(dma_time, cube_time, vec_time)
    return dma_time, cube_time, vec_time, total_time

total_cube = total_dma = total_vec = 0
summary_op_tiu_bdc = defaultdict(lambda : {"dma":0,
                                           "cube_ops":0,
                                           "vec_ops":0,
                                           "dma_time":0,
                                           "cube_time":0,
                                           "vec_time":0,
                                           "total_time":0})

ops_info = []
each_op_info = {"name": "", "dma": 0, "cube": 0, "vec": 0, "dma_time": 0, "cube_time": 0, "vec_time": 0, "total_time": 0}

theory_total_time = 0

for each_op in location:
    if each_op['core_id'] == 0 or each_op["is_local"]:
        op = each_op
        info = get_op_shape_info(op)
        dma, cube_ops, vec_ops = calc_op_dma_ops(each_op)
        dma_time, cube_time, vec_time, total_time = calc_time(dma, cube_ops, vec_ops)
        ops_info.append({"name": each_op['opcode'], "dma": dma, "cube": cube_ops, "vec": vec_ops, "dma_time": dma_time,
                         "cube_time": cube_time, "vec_time": vec_time, "total_time": total_time,
                         "info": info})
        each_op['dma']      = dma
        each_op['cube_ops'] = cube_ops
        each_op['vec_ops']  = vec_ops
        total_cube         += cube_ops
        total_dma          += dma
        total_vec          += vec_ops
        theory_total_time  += total_time
        summary_op_tiu_bdc[op['opcode']]["dma"]        += dma
        summary_op_tiu_bdc[op['opcode']]["cube_ops"]   += cube_ops
        summary_op_tiu_bdc[op['opcode']]["vec_ops"]    += vec_ops
        summary_op_tiu_bdc[op['opcode']]["dma_time"]   += dma_time
        summary_op_tiu_bdc[op['opcode']]["cube_time"]  += cube_time
        summary_op_tiu_bdc[op['opcode']]["vec_time"]   += vec_time
        summary_op_tiu_bdc[op['opcode']]["total_time"] += total_time


ops_info_pd = pd.DataFrame(ops_info)
ops_pd = pd.DataFrame(summary_op_tiu_bdc).T
ops_pd['precent']      = round(ops_pd['total_time'] / theory_total_time * 100, 2)
ops_info_pd['precent'] = round(ops_info_pd['total_time'] / theory_total_time * 100, 2)
# sort by precent
ops_pd = ops_pd.sort_values(by="precent", ascending=False)
ops_info_pd = ops_info_pd.sort_values(by="precent", ascending=False)
# add summary
ops_pd.loc["Summary"]      = ops_pd.sum()
# avoid info column to be sum
ops_info_pd.loc["Summary"] = ops_info_pd.select_dtypes(include=[float, int]).sum()
# write pcie


annotation = ["下面是每一类op汇总的情况"]
annotation += [ "理论去算的每一个op的计算量。这个op是final mlir里面的op，没有尝试去遍历fxgraph获取。\n所以搬运量都是op在final 里面的最少搬运量，没有考虑l2mem" ]
annotation.append("注意 在算的时候，出于简单，没有考虑fp16和fp32，所有的tiu耗时都是用fp16 对应engine峰值算力去算的.")
annotation.append("单位： DMA: B, Cube: ops, Vec: ops, Time: us")
annotation.append(f"DMA带宽 {CONFIG['dma']} B/us, Cube带宽 {CONFIG['cube']} ops/us, Vec带宽 {CONFIG['vec']} ops/us")
ws = wb.create_sheet(title=f"Theory Summary Op {name_info} summary")
ws.title   = f"Theory Each Op {name_info} summary"
last_row = write_to_sheet(ws, annotation, 3, 3)
last_row = write_to_sheet(ws, ops_pd, last_row + 3, 3)

last_row = write_to_sheet(ws, ["下面是每一个op的理论计算量"], last_row + 3, 3)
last_row = write_to_sheet(ws, ["info是算子的输入和输出"], last_row + 1, 3)
last_row = write_to_sheet(ws, ops_info_pd, last_row + 3, 3)

wb.save(output_path)
